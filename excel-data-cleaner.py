# This is used for large spreadsheets to copy images listed in an excel refrencing a keyword into a new folder. For support go to my github page
# github.com/

# Module for reading excel sheets https://github.com/chronossc/openpyxl
import openpyxl
# Module to copy files
from shutil import copyfile
import os,sys



############# ___ CONFIGURATION ___ ###########
# Path of your folder
path = "X:\Matlab\data"
# Name of file
file_name = "\Data_Entry_2017.xlsx"
# disease to search for
disease = "Atelectasis"
# Destination Folder
path2 = path + "\\" + disease 
# Column 1(Which contains the text)
column_text = 2
# Column 2(Which contains the images)
column_image = 1
# Last index of Columns
last_index = 112121
#Path where the images are located. Images must be in the same file
pathofimages = "X:\Matlab\data\ImageList"

############# ___ ++++++++++++++ ___ ###########



# This function finds the keywords inside the worksheets
# @keyword_ : this variable takes in a string to find inside the worksheet.
def findKeyWord(keyword_):
    # Initialising the arrays
    index = {}
    images = {}
    name_list = {}
    # Counter for storing elements inside the array
    j = 0
    for i in range(1,last_index):
        # If it contains the keyword AND does not contain the verticle line to ensure only a single disease is copied
        if( ((spreadsheet.cell(row=i,column = 2).value).find('|') == -1) and ((spreadsheet.cell(row=i,column = 2).value).find(keyword_) != -1)):
            # Storing the index of the excel spreadsheet. For testing
            index[j] = i
            # Storing image names
            images[j] = spreadsheet.cell(row=i,column = 1).value
            # Storing the description of the images
            name_list[j] = spreadsheet.cell(row=i,column = 2).value
            # Incrementing the pointer to point to the next available position in the array
            j = j + 1
    # Returning multiple variables
    return index, images, name_list


# openpyxl object
obj = openpyxl.load_workbook(path + "" + file_name)
#Get active workbook
spreadsheet = obj.active
# Getting all the information from the findKeyWord function
index, images,name_list = findKeyWord(disease)

# Printing all the images + descriptions
for i in range(0,(images.__len__() -1)):
    print(name_list[i] + " : " + images[i])
    #Try and catch to create the folder if the path (variable path 2) does not exist
    try:
        # Attempts to copy files. If FileNotFoundError exception is raised move to the except block
        copyfile(pathofimages + "\\" +images[i] , path2 + "\\" + images[i])
        print("Files copied")
    except FileNotFoundError:
        # Creating the path.
        os.mkdir(path2)
        # Copying files.
        copyfile(pathofimages + "\\" +images[i] , path2 + "\\" + images[i])
        print("Files copied")
